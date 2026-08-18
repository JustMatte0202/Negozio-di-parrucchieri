"""
Crea il file `registro_giornaliero.xlsx`: il quaderno del negozio in versione
digitale, da compilare ogni sera (anche da telefono, con Google Fogli).

Il listino dei prodotti viene letto da `listino_prodotti.csv`, così per
aggiungere prodotti basta aggiungere righe a quel file e rilanciare:

    python crea_registro.py

I servizi (taglio, barba, ecc.) sono ancora quelli di esempio: vanno
sostituiti con i prezzi veri del negozio. Nel foglio sono evidenziati in
giallo proprio per ricordarlo.
"""

import csv
from datetime import date, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CARTELLA = Path(__file__).parent
ARIAL = "Arial"
BLU_SCURO = "1F4E79"
GIALLO = "FFF2CC"        # convenzione: celle da compilare/correggere
GRIGIO = "F2F2F2"        # righe prodotto (dati già confermati)

# Servizi ancora da confermare con il listino vero del negozio.
SERVIZI = [
    ("Taglio uomo", "Servizio - Taglio", 18.0, 30),
    ("Taglio + barba", "Servizio - Taglio", 26.0, 45),
    ("Barba", "Servizio - Barba", 12.0, 20),
    ("Rasatura tradizionale", "Servizio - Barba", 15.0, 30),
    ("Taglio bambino", "Servizio - Taglio", 12.0, 25),
    ("Shampoo + styling", "Servizio - Styling", 10.0, 15),
    ("Colorazione", "Servizio - Colore", 35.0, 60),
    ("Trattamento cute", "Servizio - Trattamenti", 22.0, 30),
]


def carica_prodotti() -> list[tuple[str, str, float, int]]:
    """Legge listino_prodotti.csv -> (nome con formato, categoria, prezzo, 0)."""
    percorso = CARTELLA / "listino_prodotti.csv"
    if not percorso.exists():
        return []
    prodotti = []
    with open(percorso, encoding="utf-8") as f:
        for riga in csv.DictReader(f):
            nome = f"{riga['nome'].strip()} {riga['formato'].strip()}".strip()
            prodotti.append((nome, riga["categoria"].strip(),
                             float(riga["prezzo"]), 0))
    return prodotti


def intestazione(ws, colonne, larghezze) -> None:
    for col, (nome, largh) in enumerate(zip(colonne, larghezze), start=1):
        c = ws.cell(row=1, column=col, value=nome)
        c.font = Font(name=ARIAL, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BLU_SCURO)
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = largh
    ws.freeze_panes = "A2"


def main() -> None:
    prodotti = carica_prodotti()
    voci = [(n, c, p, d) for n, c, p, d in SERVIZI] + prodotti

    # I nomi finiscono nel menu a tendina e collegano Registro e Listino:
    # se due voci avessero lo stesso nome non si capirebbe quale è stata venduta.
    duplicati = {n for n, *_ in voci if [x[0] for x in voci].count(n) > 1}
    if duplicati:
        raise SystemExit(f"Nomi duplicati nel listino: {sorted(duplicati)}")

    wb = Workbook()

    # ------------------------------------------------------------ Istruzioni
    ws = wb.active
    ws.title = "Istruzioni"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100
    righe = [
        ("💈 Registro giornaliero del negozio", 16, True),
        ("", 11, False),
        ("Come si usa (5 minuti a fine giornata):", 12, True),
        ("1. Vai nel foglio 'Registro' e aggiungi una riga per ogni cliente servito oggi.", 11, False),
        ("2. Una riga anche per ogni PRODOTTO venduto (scegli il prodotto nella colonna Servizio).", 11, False),
        ("3. Nelle colonne Servizio, Metodo pagamento e Stato NON devi scrivere: clicca la cella e scegli dal menu a tendina.", 11, False),
        ("4. Il Prezzo è quello incassato davvero (se hai fatto uno sconto, scrivi il prezzo scontato).", 11, False),
        ("5. La colonna Cliente è FACOLTATIVA: se la compili, usa sempre lo stesso nome per la stessa persona (es. sempre 'Marco Rossi', non una volta 'Marco' e una 'Rossi Marco').", 11, False),
        ("6. La prima riga del Registro è un ESEMPIO: cancellala quando inserisci i dati veri.", 11, False),
        ("", 11, False),
        ("Il foglio 'Listino'", 12, True),
        ("Contiene tutto ciò che si può vendere, in un unico elenco ordinato:", 11, False),
        ("prima i SERVIZI (categoria che inizia per 'Servizio - '), poi i PRODOTTI (categoria 'Prodotto - ...').", 11, False),
        ("Ogni formato del prodotto è una riga a sé, perché ha un prezzo diverso (es. 250 ml e 1000 ml).", 11, False),
        ("", 11, False),
        ("ATTENZIONE: le righe dei servizi sono evidenziate in GIALLO perché sono ancora quelle di esempio.", 11, False),
        ("Vanno sostituite con i servizi e i prezzi veri del negozio.", 11, False),
        ("", 11, False),
        ("Privacy: se compili i nomi dei clienti, questo file contiene dati personali.", 11, False),
        ("Tienilo solo tra te e papà: non caricarlo mai su un sito o un repository pubblico.", 11, False),
    ]
    for i, (testo, dim, grassetto) in enumerate(righe, start=2):
        c = ws.cell(row=i, column=2, value=testo)
        c.font = Font(name=ARIAL, size=dim, bold=grassetto,
                      color=BLU_SCURO if grassetto else "000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # --------------------------------------------------------------- Listino
    ws_l = wb.create_sheet("Listino")
    ws_l.sheet_view.showGridLines = False
    intestazione(ws_l,
                 ["id_servizio", "nome_servizio", "categoria", "prezzo_listino", "durata_min"],
                 [11, 38, 30, 15, 12])
    for r, (nome, categoria, prezzo, durata) in enumerate(voci, start=2):
        e_servizio = categoria.startswith("Servizio")
        valori = [r - 1, nome, categoria, prezzo, durata]
        for col, valore in enumerate(valori, start=1):
            c = ws_l.cell(row=r, column=col, value=valore)
            c.font = Font(name=ARIAL)
            c.fill = PatternFill("solid", fgColor=GIALLO if e_servizio else GRIGIO)
            if col == 4:
                c.number_format = '"€" #,##0.00'

    # --------------------------------------------------------------- Registro
    ws_r = wb.create_sheet("Registro")
    intestazione(ws_r,
                 ["Data", "Ora", "Servizio", "Prezzo incassato", "Metodo pagamento",
                  "Stato", "Cliente (facoltativo)", "Note"],
                 [12, 8, 30, 16, 17, 13, 24, 28])

    esempio = [date(2026, 8, 18), time(17, 30), "Taglio + barba", 26.0,
               "Carta", "Completato", "Marco Rossi", "riga di ESEMPIO: cancellami"]
    for col, valore in enumerate(esempio, start=1):
        c = ws_r.cell(row=2, column=col, value=valore)
        c.font = Font(name=ARIAL)
        c.fill = PatternFill("solid", fgColor=GIALLO)

    for r in range(2, 1001):
        ws_r.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws_r.cell(row=r, column=2).number_format = "HH:MM"
        ws_r.cell(row=r, column=4).number_format = '"€" #,##0.00'

    # Il menu a tendina punta esattamente alle righe usate del Listino:
    # un riferimento diretto (non un nome definito) sopravvive meglio
    # al passaggio in Google Fogli.
    ultima = len(voci) + 1
    dv_servizio = DataValidation(
        type="list", formula1=f"=Listino!$B$2:$B${ultima}", allow_blank=True,
        showErrorMessage=True,
        error="Scegli una voce dal menu, oppure aggiungila prima nel foglio Listino.")
    dv_pagamento = DataValidation(type="list",
                                  formula1='"Contanti,Carta,Satispay,Altro"', allow_blank=True)
    dv_stato = DataValidation(type="list",
                              formula1='"Completato,No-show,Cancellato"', allow_blank=True)
    for dv, celle in ((dv_servizio, "C2:C1000"), (dv_pagamento, "E2:E1000"), (dv_stato, "F2:F1000")):
        ws_r.add_data_validation(dv)
        dv.add(celle)

    wb.save(CARTELLA / "registro_giornaliero.xlsx")
    n_servizi = sum(1 for _, c, _, _ in voci if c.startswith("Servizio"))
    print(f"registro_giornaliero.xlsx creato: {n_servizi} servizi + "
          f"{len(voci) - n_servizi} prodotti = {len(voci)} voci nel Listino.")


if __name__ == "__main__":
    main()
