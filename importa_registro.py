"""
Importa i dati veri dal registro Excel alla dashboard.

Legge `registro_giornaliero.xlsx` (il file che si compila ogni sera) e
produce i tre CSV in `data/` nello stesso formato del dataset sintetico,
così la dashboard funziona senza nessuna modifica.

Uso:
    python importa_registro.py                # scrive in data/
    python importa_registro.py --output prova # scrive in prova/ (per test)

ATTENZIONE: sovrascrive i CSV in `data/` (quindi anche il dataset sintetico
di allenamento — se vuoi tenerlo, rigeneralo con `python genera_dataset.py`).
"""

import argparse
import sys
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

FILE_REGISTRO = Path(__file__).parent / "registro_giornaliero.xlsx"


def normalizza(testo) -> str:
    return str(testo).strip().lower() if testo is not None else ""


def leggi_ora(valore) -> str:
    """Riporta l'ora nel formato HH:MM:SS usato dai CSV."""
    if isinstance(valore, time):
        return valore.strftime("%H:%M:%S")
    if isinstance(valore, datetime):
        return valore.time().strftime("%H:%M:%S")
    if valore:
        testo = str(valore).strip()
        for formato in ("%H:%M:%S", "%H:%M", "%H.%M"):
            try:
                return datetime.strptime(testo, formato).strftime("%H:%M:%S")
            except ValueError:
                pass
    return "12:00:00"  # ora mancante: mezzogiorno come segnaposto


def leggi_data(valore, n_riga: int, errori: list[str]):
    if isinstance(valore, datetime):
        return valore.date()
    if isinstance(valore, date):
        return valore
    if valore:
        testo = str(valore).strip()
        for formato in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(testo, formato).date()
            except ValueError:
                pass
    errori.append(f"  - riga {n_riga}: data mancante o non riconosciuta ({valore!r})")
    return None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="data",
                        help="cartella di destinazione dei CSV (default: data)")
    args = parser.parse_args()
    cartella = Path(__file__).parent / args.output
    cartella.mkdir(exist_ok=True)

    if not FILE_REGISTRO.exists():
        sys.exit(f"Non trovo {FILE_REGISTRO.name} nella cartella del progetto.")

    wb = load_workbook(FILE_REGISTRO, data_only=True)
    if "Listino" not in wb.sheetnames or "Registro" not in wb.sheetnames:
        sys.exit("Il file deve contenere i fogli 'Listino' e 'Registro'.")

    # ------------------------------------------------------------- Listino
    servizi = []           # righe per servizi.csv
    id_per_nome = {}       # "taglio uomo" -> 1
    for riga in wb["Listino"].iter_rows(min_row=2, max_col=5, values_only=True):
        _, nome, categoria, prezzo, durata = riga
        if not nome or normalizza(nome) in id_per_nome:
            continue
        if not isinstance(prezzo, (int, float)):
            continue  # righe di nota in fondo al foglio
        nuovo_id = len(servizi) + 1
        servizi.append({
            "id_servizio": nuovo_id,
            "nome_servizio": str(nome).strip(),
            "categoria": str(categoria).strip() if categoria else "Altro",
            "prezzo_listino": float(prezzo),
            "durata_min": int(durata) if durata is not None else 0,
        })
        id_per_nome[normalizza(nome)] = nuovo_id
    if not servizi:
        sys.exit("Il foglio 'Listino' è vuoto: compilalo prima di importare.")

    # ------------------------------------------------------------ Registro
    appuntamenti = []
    clienti = {}           # nome normalizzato -> dati cliente
    errori: list[str] = []
    avvisi: list[str] = []

    for n_riga, riga in enumerate(
        wb["Registro"].iter_rows(min_row=2, max_col=8, values_only=True), start=2
    ):
        valori = list(riga) + [None] * (8 - len(riga))
        v_data, v_ora, v_servizio, v_prezzo, v_pagamento, v_stato, v_cliente, v_note = valori

        if all(v is None or str(v).strip() == "" for v in valori):
            continue  # riga vuota
        if "esempio" in normalizza(v_note):
            avvisi.append(f"  - riga {n_riga}: riga di esempio, saltata")
            continue

        giorno = leggi_data(v_data, n_riga, errori)
        if giorno is None:
            continue

        chiave_servizio = normalizza(v_servizio)
        if chiave_servizio not in id_per_nome:
            errori.append(f"  - riga {n_riga}: servizio {v_servizio!r} non presente nel Listino")
            continue
        id_servizio = id_per_nome[chiave_servizio]

        stato = str(v_stato).strip().capitalize() if v_stato else "Completato"
        if stato not in ("Completato", "No-show", "Cancellato"):
            stato = "Completato"

        if stato != "Completato":
            prezzo = 0.0
        elif isinstance(v_prezzo, (int, float)):
            prezzo = float(v_prezzo)
        else:
            prezzo = servizi[id_servizio - 1]["prezzo_listino"]
            avvisi.append(f"  - riga {n_riga}: prezzo mancante, uso il listino (€{prezzo})")

        nome_cliente = str(v_cliente).strip() if v_cliente else "Cliente non registrato"
        chiave_cliente = normalizza(nome_cliente)
        if chiave_cliente not in clienti:
            parti = nome_cliente.split(maxsplit=1)
            clienti[chiave_cliente] = {
                "id_cliente": len(clienti) + 1,
                "nome": parti[0],
                "cognome": parti[1] if len(parti) > 1 else "",
                "anno_nascita": "",
                "canale_acquisizione": "Non registrato",
                "data_prima_visita": giorno,
            }
        cliente = clienti[chiave_cliente]
        cliente["data_prima_visita"] = min(cliente["data_prima_visita"], giorno)

        appuntamenti.append({
            "id_cliente": cliente["id_cliente"],
            "id_servizio": id_servizio,
            "data": giorno,
            "ora": leggi_ora(v_ora),
            "stato": stato,
            "prezzo_pagato": prezzo,
            "metodo_pagamento": str(v_pagamento).strip() if v_pagamento and stato == "Completato" else "",
        })

    if errori:
        print("ERRORI: correggi queste righe nel registro e rilancia l'importazione.")
        print("\n".join(errori))
        sys.exit(1)
    if not appuntamenti:
        sys.exit("Nessuna riga di dati trovata nel foglio 'Registro'.")

    appuntamenti.sort(key=lambda a: (a["data"], a["ora"]))
    for nuovo_id, app in enumerate(appuntamenti, start=1):
        app["id_appuntamento"] = nuovo_id

    # ------------------------------------------------------------- Scrittura
    import csv

    def scrivi(nome_file, righe, colonne):
        with open(cartella / nome_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=colonne)
            writer.writeheader()
            writer.writerows({c: r[c] for c in colonne} for r in righe)

    scrivi("servizi.csv", servizi,
           ["id_servizio", "nome_servizio", "categoria", "prezzo_listino", "durata_min"])
    scrivi("clienti.csv", list(clienti.values()),
           ["id_cliente", "nome", "cognome", "anno_nascita",
            "canale_acquisizione", "data_prima_visita"])
    scrivi("appuntamenti.csv", appuntamenti,
           ["id_appuntamento", "id_cliente", "id_servizio", "data", "ora",
            "stato", "prezzo_pagato", "metodo_pagamento"])

    if avvisi:
        print("Avvisi:")
        print("\n".join(avvisi))
    incasso = sum(a["prezzo_pagato"] for a in appuntamenti)
    print(f"Importati {len(appuntamenti)} appuntamenti e {len(clienti)} clienti "
          f"in '{cartella.name}/' (incasso totale €{incasso:,.2f}).")
    print("Ora avvia la dashboard con:  streamlit run dashboard.py")


if __name__ == "__main__":
    main()
